from textwrap import fill
import os
import tkinter as tk
import openpyxl as pyxl
import pandas as pd 
from PyPDF2 import PdfFileMerger
from tkinter import simpledialog

def createSolution():
    title = simpledialog.askstring(title="Name your Solution", prompt="Enter the Name for your new Solution folder")
    newSolutionPath = os.getcwd() + '\\Solution ' + title
    neededFolders = ['\\Data Files', '\\Map Data', '\\Report Files\\Accounts by route', '\\Report Files\\All Days', '\\Report Files\\Five Days']
    for folder in neededFolders:
        path = newSolutionPath + folder
        os.makedirs(path)
        print(path)

def exportMapData():
    wb = pyxl.load_workbook('Reroute Control File.xlsm') 
    solution = pd.read_excel('Reroute Control File.xlsm', 'Solution Unique stops only').dropna(0, subset=['Cu_ID'])
    rts = solution['New Rt'].unique().astype(int)
    columnsToExport = ("Acct #", "Cust #", "Week #", "New Rt", "New Delivery Day", "New Stop", "Customer Name", "DisplayAddressFull", "X_Longitude", "Y_Latitude")
    for rt in rts:
        mask = '`New Rt` == ' + str(rt)
        fileName = 'Solution ' + wb['Control'].cell(1,2).internal_value + '\\Map Data\\' 'Rt ' + str(rt) + '.tsv'
        solution.query(mask, False).to_csv(fileName,'\t', columns= columnsToExport, quoting= 1)


def mergePDF():
 
    wb = pyxl.load_workbook('Reroute Control File.xlsm') 
    solution = wb['Control'].cell(1,2).internal_value
    neededFolders = [ '\\Solution '+ solution + '\\Report Files\\Accounts by route', '\\Solution '+ solution + '\\Report Files\\All Days', '\\Solution '+ solution + '\\Report Files\\Five Days']
    cwd = os.getcwd() 
    # solution = cwd.rsplit('\\', 1)[1]

    for folder in neededFolders:
        merger = PdfFileMerger()
        os.chdir(cwd + folder)
        x = [a for a in os.listdir() if a.endswith(".pdf")]
        for pdf in x:
            merger.append(open(pdf, 'rb'))
        os.chdir(cwd)
        filename = cwd + folder +"\\Solution " + solution + ' ' + folder.rsplit('\\', 1)[1] + '.pdf'
        with open(filename, "wb") as fout:
            merger.write(fout)
        merger.close()

tk.Button(text="Create new Solution folder",command=createSolution).pack(fill=tk.X)
tk.Button(text="Create Map Data for active solution",command=exportMapData).pack(fill=tk.X)
tk.Button(text="Create merged report files for active solution",command=mergePDF).pack(fill=tk.X)

tk.mainloop()